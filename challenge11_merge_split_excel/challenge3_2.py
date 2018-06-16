#!/usr/bin/env python3
# coding=gbk

import datetime #����ʱ���������
from openpyxl import load_workbook #�����������ݱ�
from openpyxl import Workbook #�����µ����ݱ�


#��ȡ�ļ��Լ���Ӧ��
wb = load_workbook('courses.xlsx')
student_sheet = wb['students']
time_sheet = wb['time']


def combine():
    #����combine��
    combine_sheet = wb.create_sheet(title='combine')
    combine_sheet.append(['����ʱ��', '�γ�����', 'ѧϰ����', 'ѧϰʱ��'])
    # �ϲ������������ӵ� combine����
    for stu in student_sheet.values:
        #ȥ��������ͷ��һ��
        if stu[2] != 'ѧϰ����':
            #����ƥ��ʱ������
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    combine_sheet.append(list(stu) + [time[2]])
    #���Ǳ���course.xlsx
    wb.save('courses.xlsx')


def split():
    combine_sheet = wb['combine']
    #�洢combine ���е����
    split_name = []
    #������ȡ��Ӧ���
    for item in combine_sheet.values:
        if item[0] != '����ʱ��':
            split_name.append(item[0].strftime("%Y"))

    #�ֱ�洢����
    for name in set(split_name):
        #�����ļ�
        wb_temp = Workbook()
        #ɾ�����е�Ĭ�� Sheet��
        wb_temp.remove(wb_temp.active)
        #������Ӧ��������ı�
        ws = wb_temp.create_sheet(title=name)
        #д��������ݵ�����
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '����ʱ��':
                if item_by_year[0].strftime("%Y") == name:
                    ws.append(item_by_year)
        #�洢��Ӧ��ݵ������ļ�
        wb_temp.save('{}.xlsx'.format(name))


if __name__ == '__main__':
    combine()
    split()



